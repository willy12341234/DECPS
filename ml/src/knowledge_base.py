"""
RAG knowledge base: parses docx + PDF + xlsx files from ml/rag_kb/ into chunks,
builds embeddings, retrieves relevant context for chat queries.
"""
import os, re, json, numpy as np

# SentenceTransformer is optional — lazy-import to avoid HF download hang at top level
_HAVE_ST = False
_ST_MODEL = None

_PRJ = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")

_RAG_DIR = os.path.join(_PRJ, "rag_kb")

class KnowledgeBase:
    def __init__(self, model_name='paraphrase-multilingual-MiniLM-L12-v2'):
        global _HAVE_ST, _ST_MODEL
        self.model = _ST_MODEL
        if not _HAVE_ST:
            import threading
            result = [None, None]
            def _try_import():
                try:
                    import os as _os
                    _os.environ['HF_HUB_OFFLINE'] = '1'
                    from sentence_transformers import SentenceTransformer as ST
                    m = ST(model_name)
                    result[0] = m
                except Exception as e:
                    result[1] = str(e)
            t = threading.Thread(target=_try_import, daemon=True)
            t.start()
            t.join(timeout=20)
            if t.is_alive():
                print("  SentenceTransformer import timed out, using fallback matching")
            elif result[0] is not None:
                _ST_MODEL = result[0]
                _HAVE_ST = True
                self.model = _ST_MODEL
                print(f"  SentenceTransformer '{model_name}' loaded OK")
            else:
                print(f"  SentenceTransformer not available: {result[1]}, using fallback matching")
        if not _HAVE_ST:
            print("  Using fallback keyword matching (no SentenceTransformer)")
        self.chunks = []
        self.embeddings = None
        self._equipment_text = None
        self._docs_loaded = False  # PDF/DOCX lazy flag
        self._load_xlsx_only()

    def _parse_xlsx_equipment(self, path):
        """Parse equipment Excel file — extract active (not 停用/报废) instruments."""
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        chunks = []
        source = os.path.basename(path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(min_row=3, values_only=True))

            active_items = []
            for row in rows:
                if not row or not row[2]:
                    continue
                name = str(row[2]).strip() if row[2] else ''
                if not name:
                    continue

                model = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                location = str(row[7]).strip() if len(row) > 7 and row[7] else ''
                status = str(row[10]).strip() if len(row) > 10 and row[10] else ''
                dept = str(row[6]).strip() if len(row) > 6 and row[6] else ''
                asset_id = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                note = str(row[8]).strip() if len(row) > 8 and row[8] else ''

                # Only exclude explicitly retired/scrapped equipment
                if status in ('decommissioned', 'scrapped'):
                    continue

                active_items.append({
                    'name': name, 'model': model, 'location': location,
                    'dept': dept, 'asset_id': asset_id, 'note': note,
                })

            if not active_items:
                continue

            category = sheet_name
            lines = [f"【公司设备清单 - {category}】"]
            for item in active_items:
                parts = [f"{item['name']}"]
                if item['model']:
                    parts.append(f"型号: {item['model']}")
                if item['location']:
                    parts.append(f"位置: {item['location']}")
                if item['note']:
                    parts.append(f"备注: {item['note']}")
                lines.append(' - ' + ', '.join(parts))

            text = '\n'.join(lines)
            chunks.append({
                'text': text,
                'section': f'设备台账 - {category}',
                'mechanism': f'可用设备-{category}',
                'type': 'content',
                'source': source,
                'equipment_category': category,
                'equipment_list': active_items,
            })

        wb.close()
        return chunks

    def _parse_docx(self, path):
        from docx import Document
        doc = Document(path)
        chunks = []
        current_section = "概述"
        current_mechanism = ""
        buffer = []
        for para in doc.paragraphs:
            t = para.text.strip()
            if not t:
                continue
            if re.match(r'^第[一二三四五六七八九十]+章\s+', t):
                if buffer:
                    chunks.append({'text': '\n'.join(buffer), 'section': current_section,
                                   'mechanism': current_mechanism, 'type': 'content',
                                   'source': os.path.basename(path)})
                    buffer = []
                current_section = t[:40]
                current_mechanism = ""
                chunks.append({'text': t, 'section': current_section,
                               'mechanism': '', 'type': 'header',
                               'source': os.path.basename(path)})
                continue
            if re.match(r'^\d+\.\d+\s+', t):
                if buffer:
                    chunks.append({'text': '\n'.join(buffer), 'section': current_section,
                                   'mechanism': current_mechanism, 'type': 'content',
                                   'source': os.path.basename(path)})
                    buffer = []
                m = re.match(r'\d+\.\d+\s+(.*)', t)
                title = m.group(1).strip() if m else t
                if '反应' in title or '机制' in title or '规则' in title:
                    current_mechanism = title[:30]
                chunks.append({'text': t, 'section': current_section,
                               'mechanism': current_mechanism, 'type': 'subheader',
                               'source': os.path.basename(path)})
                continue
            buffer.append(t)
        if buffer:
            chunks.append({'text': '\n'.join(buffer), 'section': current_section,
                           'mechanism': current_mechanism, 'type': 'content',
                           'source': os.path.basename(path)})
        return chunks

    def _parse_pdf(self, path):
        from pypdf import PdfReader
        reader = PdfReader(path)
        chunks = []
        source = os.path.basename(path)
        for i, page in enumerate(reader.pages):
            text = page.extract_text()
            if not text.strip():
                continue
            chunks.append({
                'text': text.strip(),
                'section': f'{source} - 第{i+1}页',
                'mechanism': source.replace('.pdf', '').replace('-', ' ')[:60],
                'type': 'content',
                'source': source
            })
        return chunks

    def _load_all_documents(self):
        chunks = []

        # 1. Load all .docx files from rag_kb/
        if os.path.isdir(_RAG_DIR):
            for fname in sorted(os.listdir(_RAG_DIR)):
                if fname.lower().endswith('.docx'):
                    fpath = os.path.join(_RAG_DIR, fname)
                    try:
                        docx_chunks = self._parse_docx(fpath)
                        chunks.extend(docx_chunks)
                        print(f"  Loaded DOCX: {fname} ({len(docx_chunks)} chunks)")
                    except Exception as e:
                        print(f"  Skip DOCX {fname}: {e}")
        else:
            print("  No RAG knowledge base directory found (ml/rag_kb/)")

        # 2. Load all PDFs from rag_kb/
        if os.path.isdir(_RAG_DIR):
            for fname in sorted(os.listdir(_RAG_DIR)):
                if fname.lower().endswith('.pdf'):
                    fpath = os.path.join(_RAG_DIR, fname)
                    try:
                        pdf_chunks = self._parse_pdf(fpath)
                        chunks.extend(pdf_chunks)
                        print(f"  Loaded PDF: {fname} ({len(pdf_chunks)} chunks)")
                    except Exception as e:
                        print(f"  Skip PDF {fname}: {e}")

        # 3. Load equipment XLSX from rag_kb/
        if os.path.isdir(_RAG_DIR):
            for fname in sorted(os.listdir(_RAG_DIR)):
                if fname.lower().endswith('.xlsx'):
                    fpath = os.path.join(_RAG_DIR, fname)
                    try:
                        xlsx_chunks = self._parse_xlsx_equipment(fpath)
                        chunks.extend(xlsx_chunks)
                        print(f"  Loaded XLSX: {fname} ({len(xlsx_chunks)} chunks)")
                    except Exception as e:
                        print(f"  Skip XLSX {fname}: {e}")

        return chunks

    def _load_xlsx_only(self):
        """Fast init: only load equipment XLSX. PDF/DOCX loaded lazily on first retrieve()."""
        if not os.path.isdir(_RAG_DIR):
            print("No RAG database directory found")
            return
        for fname in sorted(os.listdir(_RAG_DIR)):
            if fname.lower().endswith('.xlsx'):
                fpath = os.path.join(_RAG_DIR, fname)
                try:
                    xlsx_chunks = self._parse_xlsx_equipment(fpath)
                    self.chunks.extend(xlsx_chunks)
                    print(f"  Loaded XLSX: {fname} ({len(xlsx_chunks)} chunks)")
                except Exception as e:
                    print(f"  Skip XLSX {fname}: {e}")

    def _ensure_documents_loaded(self):
        """Lazy-load all PDFs and DOCX files if not already loaded."""
        if self._docs_loaded or not os.path.isdir(_RAG_DIR):
            return
        self._docs_loaded = True
        raw = []
        for fname in sorted(os.listdir(_RAG_DIR)):
            fpath = os.path.join(_RAG_DIR, fname)
            if fname.lower().endswith('.docx'):
                try:
                    docx_chunks = self._parse_docx(fpath)
                    raw.extend(docx_chunks)
                    print(f"  Loaded DOCX: {fname} ({len(docx_chunks)} chunks)")
                except Exception as e:
                    print(f"  Skip DOCX {fname}: {e}")
            elif fname.lower().endswith('.pdf'):
                try:
                    pdf_chunks = self._parse_pdf(fpath)
                    raw.extend(pdf_chunks)
                    print(f"  Loaded PDF: {fname} ({len(pdf_chunks)} chunks)")
                except Exception as e:
                    print(f"  Skip PDF {fname}: {e}")

        if not raw:
            print("No documents found for lazy load, using fallback rules")
            self.chunks = self._fallback_rules() + self.chunks
            return

        new_chunks = []
        for c in raw:
            if c['type'] == 'content' and len(c['text']) > 600:
                sentences = re.split(r'(?<=[。；\n])\s*', c['text'])
                part = []
                p_len = 0
                for s in sentences:
                    if p_len + len(s) > 600 and part:
                        new_chunks.append({**c, 'text': ''.join(part)})
                        part = [s]; p_len = len(s)
                    else:
                        part.append(s); p_len += len(s)
                if part:
                    new_chunks.append({**c, 'text': ''.join(part)})
            else:
                new_chunks.append(c)

        # Prepend before equipment chunks (equipment stays at end)
        self.chunks = new_chunks + self.chunks

        texts = [c['text'] for c in self.chunks]
        if texts and self.model is not None:
            self.embeddings = self.model.encode(texts, show_progress_bar=False)
            self.embeddings = self.embeddings / (np.linalg.norm(self.embeddings, axis=1, keepdims=True) + 1e-10)
        print(f"Knowledge base: {len(self.chunks)} chunks from {len(set(c['source'] for c in self.chunks))} files, emb_dim={self.embeddings.shape[1] if self.embeddings is not None else 0}")

    def _fallback_rules(self):
        return [
            {'text': t, 'section': '内置规则', 'mechanism': m, 'type': 'content', 'source': 'fallback'}
            for t, m in [
                ("美拉德反应：还原糖（乳糖、葡萄糖）的羰基与API的伯胺/仲胺发生反应，生成席夫碱和类黑精色素。触发条件：高温(>40°C)、高湿、碱性pH。典型API：加巴喷丁、氨氯地平、普瑞巴林。", "美拉德反应"),
                ("氧化反应：辅料中的过氧化物（PEG、吐温80、PVP）或金属离子催化API氧化。酚类、硫醚、巯基、叔胺易感。触发条件：氧存在、光照、金属离子。", "氧化反应"),
                ("水解反应：酯键、酰胺键、内酰胺、内酯在水分存在下裂解。酸或碱催化。典型API：阿司匹林（酯）、青霉素（内酰胺）。", "水解反应"),
                ("酯化/转酯化：API的羧酸与辅料羟基（PEG、甘油）反应生成酯。酸催化。典型API：布洛芬、吲哚美辛。", "酯化反应"),
                ("光降解：UV/可见光激发API或光敏辅料（TiO₂、核黄素）产生活性氧。硝基、芳卤、烯烃易感。", "光降解"),
                ("固态酸碱反应：硬脂酸镁等碱性辅料升高微环境pH，催化酸性API水解或成盐。", "酸碱反应"),
                ("络合反应：金属离子（Ca²⁺、Mg²⁺、Fe³⁺）与API的β-二酮、儿茶酚、羧酸配位。四环素-钙螯合降低吸收。", "络合反应"),
                ("ICH Q1A(R2)加速条件：40°C±2°C/75%RH±5%RH，时间点0、3、6个月。长期条件：25°C±2°C/60%RH±5%RH。", "ICH指南"),
            ]
        ]

    def get_equipment_summary(self):
        """Return a concise markdown summary of all available equipment."""
        if self._equipment_text:
            return self._equipment_text
        equip_chunks = [c for c in self.chunks if c.get('equipment_category')]
        if not equip_chunks:
            return ''
        lines = ['## List of Available Instruments and Equipment (Only these listed instruments can be used for designing experiments)']
        for c in equip_chunks:
            cat = c["equipment_category"]
            items = c.get('equipment_list', [])
            models = {}
            for item in items:
                key = (item['name'], item['model'])
                if key not in models:
                    models[key] = []
                if item['location']:
                    if item['location'] not in models[key]:
                        models[key].append(item['location'])
            grouped = []
            for (name, model), locs in models.items():
                s = f"- **{name}**"
                if model:
                    s += f" ({model})"
                if locs:
                    s += f" @ {', '.join(locs)}"
                grouped.append(s)
            lines.append(f'\n### {cat} ({len(items)} units)')
            lines.extend(grouped)
        self._equipment_text = '\n'.join(lines)
        return self._equipment_text

    def retrieve(self, query, top_k=5):
        self._ensure_documents_loaded()
        if len(self.chunks) == 0:
            return []
        if self.embeddings is not None and self.model is not None:
            q_emb = self.model.encode([query], show_progress_bar=False)
            q_emb = q_emb / (np.linalg.norm(q_emb) + 1e-10)
            scores = self.embeddings @ q_emb.T
            scores = scores.flatten()
            top_idx = np.argsort(scores)[-top_k:][::-1]
            return [(self.chunks[i], float(scores[i])) for i in top_idx]
        # Fallback: keyword matching
        q_lower = query.lower()
        scored = []
        for c in self.chunks:
            text = c['text'].lower()
            score = text.count(q_lower) * 2
            for kw in query.split():
                if len(kw) > 1:
                    score += text.count(kw.lower())
            if score > 0:
                scored.append((c, score))
        scored.sort(key=lambda x: -x[1])
        return [(c, s / max(s for _, s in scored)) for c, s in scored[:top_k]] if scored else []

    def build_prompt(self, query, api_name=None, exc1_name=None, days=None, cond=None, model_result=None):
        retrieved = self.retrieve(query, top_k=4)
        lines = []
        if model_result:
            prob = model_result.get('prob', 0)
            level = model_result.get('level', 'unknown')
            has_imp = model_result.get('has', False)
            imp_count = model_result.get('impurity_count', None)
            total_pct = model_result.get('total_impurity_pct', None)
            max_pct = model_result.get('max_single_impurity_pct', None)
            status = "⚠ 预测会产生杂质" if has_imp else "✓ 预测相容"
            lines.append(f"【模型预测】{status}（风险等级: {level}，概率: {prob:.1%}）")
            if imp_count is not None:
                lines.append(f"  预计杂质个数: {imp_count:.1f}，总杂: {total_pct:.3f}%，最大单杂: {max_pct:.3f}%")
            if api_name and exc1_name:
                lines.append(f"  API: {api_name}，辅料: {exc1_name}")
                if cond:
                    lines.append(f"  条件: {cond}，{days}天")

        if retrieved:
            lines.append("\n【相关知识】")
            seen_mechanisms = set()
            for chunk, score in retrieved:
                mech = chunk.get('mechanism', '')
                if mech and mech not in seen_mechanisms:
                    lines.append(f"\n■ {mech}（相关度: {score:.2f}）")
                    seen_mechanisms.add(mech)
                text = chunk['text']
                if len(text) > 300:
                    text = text[:300] + '...'
                lines.append(f"  {text}")

        if not lines:
            lines.append("暂无相关检索结果。请提供更具体的API名称、辅料名称或实验条件。")

        return '\n'.join(lines)


_kb = None
def get_kb():
    global _kb
    if _kb is None:
        _kb = KnowledgeBase()
    return _kb
